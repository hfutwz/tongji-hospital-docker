#!/usr/bin/env python3
"""
离线全量训练脚本：从 Excel 直接读取数据，不依赖 MySQL，使用 openpyxl
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import json
import datetime
import re
from openpyxl import load_workbook

from app.models.statistical_model import TraumaStatisticalModel, CAUSE_NAMES, TIME_PERIOD_NAMES, SEASON_NAMES
from app.models.feature_builder import _normalize_cause, _resolve_district

EXCEL_PATH = "../trauma_patient_data.xlsx"
MODEL_DIR = "models/"

# 列索引（基于实际 Excel 结构）
COL_CAUSE = 205           # '受伤原因:'
COL_ADMISSION_DATE = 5    # '接诊日期：'
COL_ADMISSION_TIME = 6    # '接诊时间：'
COL_LOCATION = 271        # '创伤发生地...'

def calculate_time_period(time_str):
    """
    根据接诊时间计算时段
    0: 夜间(00:00-07:59), 1: 早高峰(08:00-09:59), 2: 午高峰(10:00-11:59)
    3: 下午(12:00-16:59), 4: 晚高峰(17:00-19:59), 5: 晚上(20:00-23:59)
    """
    if not time_str:
        return -1
    
    # 尝试提取时间中的小时
    # 格式可能是 "11:00"、"1100"、"11" 等
    time_str = str(time_str).strip()
    
    # 匹配 4位数字格式 (如 1100)
    if re.match(r'^\d{4}$', time_str):
        hour = int(time_str[:2])
    # 匹配 HH:MM 格式
    elif re.match(r'^(\d{1,2}):\d{2}$', time_str):
        hour = int(time_str.split(':')[0])
    # 匹配纯数字小时
    elif re.match(r'^\d{1,2}$', time_str):
        hour = int(time_str)
    else:
        return -1
    
    if 0 <= hour <= 7:
        return 0  # 夜间
    elif 8 <= hour <= 9:
        return 1  # 早高峰
    elif 10 <= hour <= 11:
        return 2  # 午高峰
    elif 12 <= hour <= 16:
        return 3  # 下午
    elif 17 <= hour <= 19:
        return 4  # 晚高峰
    elif 20 <= hour <= 23:
        return 5  # 晚上
    else:
        return -1

def calculate_season(date_val):
    """
    根据接诊日期计算季节
    0: 春季(3-5月), 1: 夏季(6-9月), 2: 秋季(10-12月), 3: 冬季(1-2月)
    """
    if not date_val:
        return -1
    
    try:
        # 尝试解析日期
        date_str = str(date_val).strip()
        
        # 格式可能是 "2024-03-15"、"2024/03/15"、"2024.03.15" 或 Excel 日期
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d']:
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                month = dt.month
                break
            except ValueError:
                continue
        else:
            # 尝试从字符串中提取年月日
            match = re.search(r'(\d{4})[-/.]?(\d{1,2})[-/.]?(\d{1,2})', date_str)
            if match:
                month = int(match.group(2))
            else:
                return -1
        
        if 3 <= month <= 5:
            return 0  # 春季
        elif 6 <= month <= 9:
            return 1  # 夏季
        elif 10 <= month <= 12:
            return 2  # 秋季
        elif 1 <= month <= 2:
            return 3  # 冬季
        else:
            return -1
    except Exception:
        return -1

def load_data_from_excel():
    """从 Excel 读取训练数据"""
    print(f"正在读取 Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    # 获取表头
    headers = [cell.value for cell in ws[1]]
    print(f"列数: {len(headers)}")
    print(f"使用列: 受伤原因={COL_CAUSE}, 接诊日期={COL_ADMISSION_DATE}, 接诊时间={COL_ADMISSION_TIME}, 创伤发生地={COL_LOCATION}")
    
    records = []
    row_count = 0
    skipped_no_time = 0
    skipped_no_cause = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        if row_count % 500 == 0:
            print(f"  已处理 {row_count} 行...")
        
        # 解析伤因
        cause_raw = row[COL_CAUSE]
        cause = _normalize_cause(cause_raw)
        if cause < 0 or cause > 4:
            skipped_no_cause += 1
            continue
        
        # 解析时段（从接诊时间）
        time_period = calculate_time_period(row[COL_ADMISSION_TIME])
        if time_period < 0:
            skipped_no_time += 1
            continue
        
        # 解析季节（从接诊日期）
        season = calculate_season(row[COL_ADMISSION_DATE])
        
        # 解析行政区
        district = _resolve_district(row[COL_LOCATION])
        
        records.append({
            'time_period': time_period,
            'season': season,
            'district': district,
            'injury_cause_category': cause,
            'injury_location': row[COL_LOCATION],
        })
    
    print(f"总记录数: {row_count}")
    print(f"跳过(无时段时间): {skipped_no_time}")
    print(f"跳过(无伤因): {skipped_no_cause}")
    print(f"有效训练样本: {len(records)}")
    return records

def evaluate_model(model, records):
    """评估模型 Top-1 准确率"""
    hits = 0
    total = 0
    for r in records:
        p = r.get('time_period', -1)
        s = r.get('season', -1)
        c = r.get('injury_cause_category')
        if p < 0 or c is None:
            continue
        pred = model.predict_by_period_season(p, s)
        predicted = max((k for k in range(5)), key=lambda k: pred.get(k, 0))
        if predicted == c:
            hits += 1
        total += 1
    
    accuracy = round(hits / total, 4) if total else 0.0
    return {'top1_accuracy': accuracy, 'sample_count': total}

def train_full():
    """执行全量训练"""
    os.makedirs(MODEL_DIR, exist_ok=True)
    
    # 加载数据
    records = load_data_from_excel()
    
    # 按时间序列切分：前80%训练，后20%测试（避免数据泄露）
    split = max(1, int(len(records) * 0.8))
    train_records = records[:split]
    test_records = records[split:]
    print(f"\n训练集: {len(train_records)} 条，测试集: {len(test_records)} 条")
    
    # 用训练集训练（用于评估）
    print("正在训练统计模型...")
    model_eval = TraumaStatisticalModel()
    model_eval.fit(train_records)
    
    # 在测试集上评估（out-of-sample）
    print("正在评估模型（测试集）...")
    metrics = evaluate_model(model_eval, test_records)
    metrics['train_count'] = len(train_records)
    metrics['test_count'] = len(test_records)
    
    # 用全量数据重训最终模型（覆盖最新数据）
    print("正在用全量数据训练最终模型...")
    model = TraumaStatisticalModel()
    model.fit(records)
    
    # 生成版本号
    version = f"v1_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}"
    
    # 保存模型文件
    versioned_pkl = os.path.join(MODEL_DIR, f"{version}.pkl")
    current_pkl = os.path.join(MODEL_DIR, "current.pkl")
    
    model.save(versioned_pkl)
    model.save(current_pkl)
    
    # 保存元数据
    meta = {
        'version': version,
        'version_num': 1,
        'trained_count': len(records),
        'delta': len(records),
        'district_count': model.meta['district_count'],
        'created_at': datetime.datetime.now().isoformat(),
        'metrics': metrics,
        'features': ['time_period', 'season', 'district', 'injury_cause_category'],
    }
    
    versioned_meta = os.path.join(MODEL_DIR, f"{version}.meta.json")
    current_meta = os.path.join(MODEL_DIR, "current.meta.json")
    
    with open(versioned_meta, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    with open(current_meta, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    
    return {
        'status': 'success',
        'version': version,
        'trained_count': len(records),
        'metrics': metrics,
        'model': model,
    }

def print_model_report(model, metrics):
    """打印模型报告"""
    print("\n" + "="*60)
    print("模型训练报告")
    print("="*60)
    
    print(f"\n【样本统计】")
    print(f"  训练样本总数: {model.meta['trained_count']}")
    print(f"  有行政区信息: {model.meta['district_count']}")
    
    print(f"\n【评估指标】")
    print(f"  Top-1 准确率: {metrics['top1_accuracy']*100:.2f}%")
    print(f"  评估样本数: {metrics['sample_count']}")
    
    print(f"\n【全局伤因分布】")
    global_dist = model._global()
    for c in range(5):
        pct = global_dist.get(c, 0) * 100
        print(f"  {CAUSE_NAMES[c]}: {pct:.1f}%")
    
    print(f"\n【时段分布】")
    for p in range(6):
        dist = model.predict_by_period(p)
        top_cause = max((c for c in range(5)), key=lambda c: dist.get(c, 0))
        print(f"  {TIME_PERIOD_NAMES[p]}: 主要伤因={CAUSE_NAMES[top_cause]}, 样本={dist.get('_sample_n', 0)}")
    
    print(f"\n【季节分布】")
    for s in range(4):
        dist = model.predict_by_season(s)
        top_cause = max((c for c in range(5)), key=lambda c: dist.get(c, 0))
        print(f"  {SEASON_NAMES[s]}: 主要伤因={CAUSE_NAMES[top_cause]}")
    
    print(f"\n【行政区覆盖】")
    district_counts = {}
    for (d, c), v in model.district_cause.items():
        district_counts[d] = district_counts.get(d, 0) + v
    for d, count in sorted(district_counts.items(), key=lambda x: -x[1])[:10]:
        print(f"  {d}: {count}条")
    
    print("\n" + "="*60)

if __name__ == "__main__":
    print("="*60)
    print("创伤预测模型 - 离线全量训练")
    print("="*60)
    
    result = train_full()
    
    if result['status'] == 'success':
        print_model_report(result['model'], result['metrics'])
        print(f"\n✅ 训练成功！")
        print(f"   版本: {result['version']}")
        print(f"   模型文件: models/current.pkl")
        print(f"   元数据: models/current.meta.json")
    else:
        print(f"\n❌ 训练失败: {result.get('reason', '未知错误')}")
        sys.exit(1)
